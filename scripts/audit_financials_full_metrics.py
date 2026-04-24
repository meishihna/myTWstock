#!/usr/bin/env python3
"""
audit_financials_full_metrics.py — 檢查 financials_store 是否同時滿足：
  - annual.periods 長度 == 8
  - quarterly.periods 長度 == 32
  - 下列 14 個 series 鍵在「年表 8 期」內皆存在且每期非 null
  - 季表：同上，但 **不檢查** 期別為 **2026-03-31**（2026 第一季）之欄位（可全 null）

鍵名與 data/financials_store 內 JSON（如 2330.json）一致。

Usage:
  python scripts/audit_financials_full_metrics.py
  python scripts/audit_financials_full_metrics.py --csv data/incomplete_full_metrics.csv
  python scripts/audit_financials_full_metrics.py --md data/financials_full_metrics_report.md
  python scripts/audit_financials_full_metrics.py --xlsx data/financials_full_metrics_report.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
from datetime import datetime, timezone

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIN_DIR = os.path.join(PROJECT_ROOT, "data", "financials_store")

REQUIRED_KEYS = (
    "Revenue",
    "Gross Profit",
    "Gross Margin (%)",
    "Selling & Marketing Exp",
    "R&D Exp",
    "General & Admin Exp",
    "Operating Income",
    "Operating Margin (%)",
    "Net Income",
    "Net Margin (%)",
    "Op Cash Flow",
    "Investing Cash Flow",
    "Financing Cash Flow",
    "CAPEX",
)

NA = 8
NQ = 32

# 季表：此季末欄位不強制要有數（例如財報尚未公告、僅 EPS 先入檔）
EXCLUDED_QUARTERLY_ENDS: frozenset[str] = frozenset({"2026-03-31"})


def _period_iso10(p) -> str:
    s = str(p).strip().split()[0]
    return s[:10] if len(s) >= 10 else s


def _bad_val(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return True
    return False


def _check_annual(block: dict | None, n_expect: int, label: str) -> list[str]:
    issues: list[str] = []
    if not block or not isinstance(block, dict):
        issues.append(f"{label}: 區塊缺失")
        return issues
    periods = block.get("periods")
    if not isinstance(periods, list):
        issues.append(f"{label}: periods 非列表")
        return issues
    if len(periods) != n_expect:
        issues.append(f"{label}: 期數 {len(periods)} != {n_expect}")

    series = block.get("series")
    if not isinstance(series, dict):
        issues.append(f"{label}: series 非物件")
        return issues

    for key in REQUIRED_KEYS:
        if key not in series:
            issues.append(f"{label}: 缺鍵 {key}")
            continue
        arr = series[key]
        if not isinstance(arr, list):
            issues.append(f"{label}: {key} 非陣列")
            continue
        if len(arr) != n_expect:
            issues.append(f"{label}: {key} 長度 {len(arr)} != {n_expect}")
            continue
        for i, v in enumerate(arr):
            if _bad_val(v):
                issues.append(f"{label}: {key}[{i}] 為空")
                break
    return issues


def _check_quarterly(block: dict | None, n_expect: int, label: str) -> list[str]:
    """季表 32 欄；期別在 EXCLUDED_QUARTERLY_ENDS 者不檢查是否為 null。"""
    issues: list[str] = []
    if not block or not isinstance(block, dict):
        issues.append(f"{label}: 區塊缺失")
        return issues
    periods = block.get("periods")
    if not isinstance(periods, list):
        issues.append(f"{label}: periods 非列表")
        return issues
    if len(periods) != n_expect:
        issues.append(f"{label}: 期數 {len(periods)} != {n_expect}")

    series = block.get("series")
    if not isinstance(series, dict):
        issues.append(f"{label}: series 非物件")
        return issues

    skip_idx = {
        i
        for i, p in enumerate(periods)
        if _period_iso10(p) in EXCLUDED_QUARTERLY_ENDS
    }

    for key in REQUIRED_KEYS:
        if key not in series:
            issues.append(f"{label}: 缺鍵 {key}")
            continue
        arr = series[key]
        if not isinstance(arr, list):
            issues.append(f"{label}: {key} 非陣列")
            continue
        if len(arr) != n_expect:
            issues.append(f"{label}: {key} 長度 {len(arr)} != {n_expect}")
            continue
        for i, v in enumerate(arr):
            if i in skip_idx:
                continue
            if _bad_val(v):
                issues.append(f"{label}: {key}[{i}] 為空")
                break
    return issues


def _write_markdown_report(
    path: str,
    *,
    total: int,
    complete_n: int,
    complete_tickers: list[str],
    incomplete: list[dict],
) -> None:
    """繁中 Markdown，方便直接閱讀或列印。"""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines: list[str] = [
        "# 財務 JSON 完整度報告",
        "",
        f"- 產生時間：{now}",
        f"- 資料目錄：`data/financials_store/`",
        "",
        "## 完整定義",
        "",
        "- **年表**：8 個年度；下列 14 個指標每期皆有值（非 null）。",
        "- **季表**：32 季；同上；**不檢查**期別 `2026-03-31`（2026 第一季）是否為空。",
        "",
        "### 檢查之 14 個欄位（JSON 鍵名）",
        "",
        ", ".join(f"`{k}`" for k in REQUIRED_KEYS),
        "",
        "## 摘要",
        "",
        "| 項目 | 檔案數 |",
        "|------|--------|",
        f"| 總計 | {total} |",
        f"| **達標** | **{complete_n}** |",
        f"| **未達標** | **{len(incomplete)}** |",
        "",
    ]

    lines += [
        "## 達標公司代號",
        "",
    ]
    ct = sorted(complete_tickers)
    if not ct:
        lines.append("（無）")
    else:
        # 每行約 12 個四位數代號，方便肉眼掃描
        row: list[str] = []
        for i, t in enumerate(ct):
            row.append(t)
            if len(row) >= 12 or i == len(ct) - 1:
                lines.append(" ".join(row))
                row = []
    lines.append("")

    lines += [
        "## 未達標清單",
        "",
        "| 代號 | 問題筆數 | 說明（首則；其餘見下方明細） |",
        "|------|----------|--------------------------------|",
    ]
    for row in incomplete:
        t = row["ticker"]
        iss: list = row["issues"]
        n = len(iss)
        first = str(iss[0]).replace("|", "\\|")
        if len(first) > 120:
            first = first[:117] + "..."
        lines.append(f"| {t} | {n} | {first} |")
    lines.append("")

    lines += [
        "---",
        "## 未達標明細（每檔完整問題列表）",
        "",
    ]
    for row in incomplete:
        t = row["ticker"]
        iss = row["issues"]
        lines.append(f"### {t}")
        lines.append("")
        for s in iss:
            lines.append(f"- {s}")
        lines.append("")

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines))


def _write_excel_report(
    path: str,
    *,
    total: int,
    complete_n: int,
    complete_tickers: list[str],
    incomplete: list[dict],
) -> None:
    """輸出 .xlsx：摘要、達標、未達標（合併說明）、未達標明細（逐條）。"""
    try:
        import pandas as pd
    except ImportError as e:
        raise RuntimeError("需要 pandas：pip install pandas") from e
    try:
        from openpyxl.styles import Alignment
    except ImportError as e:
        raise RuntimeError("需要 openpyxl：pip install openpyxl") from e

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    summary_df = pd.DataFrame(
        {
            "項目": ["總計", "達標", "未達標", "產生時間(UTC)", "資料目錄"],
            "內容": [
                str(total),
                str(complete_n),
                str(len(incomplete)),
                now,
                "data/financials_store/",
            ],
        }
    )

    rules_df = pd.DataFrame(
        {
            "規則": [
                "年表 8 期；季表 32 期；14 個指標每期非 null",
                "季表不檢查期別 2026-03-31（2026 第一季）",
            ]
        }
    )

    keys_df = pd.DataFrame({"檢查欄位（JSON 鍵）": list(REQUIRED_KEYS)})

    ok_df = pd.DataFrame({"代號": sorted(complete_tickers)})

    inc_rows: list[dict[str, str | int]] = []
    for row in incomplete:
        iss = row["issues"]
        inc_rows.append(
            {
                "代號": row["ticker"],
                "問題筆數": len(iss),
                "問題說明（全部）": "\n".join(str(x) for x in iss),
            }
        )
    inc_df = pd.DataFrame(inc_rows)

    detail_rows: list[dict[str, str | int]] = []
    for row in incomplete:
        t = row["ticker"]
        for j, s in enumerate(row["issues"], 1):
            detail_rows.append({"代號": t, "序號": j, "問題": str(s)})
    detail_df = pd.DataFrame(detail_rows)

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="摘要", index=False)
        rules_df.to_excel(writer, sheet_name="完整定義", index=False)
        keys_df.to_excel(writer, sheet_name="檢查欄位", index=False)
        ok_df.to_excel(writer, sheet_name="達標", index=False)
        inc_df.to_excel(writer, sheet_name="未達標", index=False)
        detail_df.to_excel(writer, sheet_name="未達標明細", index=False)

    # 自動換行、欄寬（未達標「問題說明」）
    from openpyxl import load_workbook

    wb = load_workbook(path)
    wrap = Alignment(wrap_text=True, vertical="top")
    if "未達標" in wb.sheetnames:
        ws = wb["未達標"]
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 100
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=3).alignment = wrap
    if "未達標明細" in wb.sheetnames:
        ws = wb["未達標明細"]
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 90
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=3).alignment = wrap
        ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from utils import setup_stdout

    setup_stdout()
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", metavar="PATH", help="輸出未達標清單 CSV")
    ap.add_argument(
        "--md",
        metavar="PATH",
        help="輸出人可讀繁中 Markdown 報告（摘要＋達標列表＋未達標表與明細）",
    )
    ap.add_argument(
        "--xlsx",
        metavar="PATH",
        help="輸出 Excel .xlsx（摘要、完整定義、檢查欄位、達標、未達標、未達標明細）",
    )
    args = ap.parse_args()

    if not os.path.isdir(FIN_DIR):
        print(f"Not found: {FIN_DIR}", file=sys.stderr)
        return 1

    incomplete: list[dict[str, str | list]] = []
    complete_tickers: list[str] = []
    complete_n = 0

    for fn in sorted(os.listdir(FIN_DIR)):
        if not fn.endswith(".json"):
            continue
        ticker = fn.replace(".json", "")
        path = os.path.join(FIN_DIR, fn)
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            incomplete.append(
                {
                    "ticker": ticker,
                    "issues": [f"JSON 錯誤: {e}"],
                }
            )
            continue

        q_issues = _check_quarterly(data.get("quarterly"), NQ, "季表")
        a_issues = _check_annual(data.get("annual"), NA, "年表")
        all_issues = q_issues + a_issues
        if all_issues:
            incomplete.append({"ticker": ticker, "issues": all_issues})
        else:
            complete_n += 1
            complete_tickers.append(ticker)

    total = complete_n + len(incomplete)
    print("=== 完整定義：8 年 + 32 季；14 鍵每期有值（季表略過 2026-03-31）===")
    print(", ".join(REQUIRED_KEYS))
    print()
    print(f"總檔案數: {total}")
    print(f"達標: {complete_n}")
    print(f"未達標: {len(incomplete)}")
    print()
    print("--- 未達標代號（前 50）---")
    for row in incomplete[:50]:
        t = row["ticker"]
        iss = row["issues"]
        print(f"  {t}: {iss[0]}" + (f" (+{len(iss)-1} 項)" if len(iss) > 1 else ""))
    if len(incomplete) > 50:
        print(f"  ... 共 {len(incomplete)} 檔，請用 --csv 看全部")

    if args.csv:
        os.makedirs(os.path.dirname(args.csv) or ".", exist_ok=True)
        with open(args.csv, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["ticker", "issue_count", "issues_joined"])
            for row in incomplete:
                iss = row["issues"]
                w.writerow([row["ticker"], len(iss), " | ".join(iss)])
        print(f"\nWrote {args.csv}")

    if args.md:
        _write_markdown_report(
            args.md,
            total=total,
            complete_n=complete_n,
            complete_tickers=complete_tickers,
            incomplete=incomplete,
        )
        print(f"Wrote {args.md}")

    if args.xlsx:
        try:
            _write_excel_report(
                args.xlsx,
                total=total,
                complete_n=complete_n,
                complete_tickers=complete_tickers,
                incomplete=incomplete,
            )
            print(f"Wrote {args.xlsx}")
        except RuntimeError as e:
            print(str(e), file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
